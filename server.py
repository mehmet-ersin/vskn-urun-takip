import json
import os
import sys
from datetime import datetime
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl

app = Flask(__name__, static_folder='public', static_url_path='')
CORS(app)
BASE = Path(__file__).parent
VERILER = BASE / "veriler"
PUBLIC = BASE / "public"

def h(val):
    if val is None: return ""
    if isinstance(val, datetime): return val.strftime("%d.%m.%Y")
    if isinstance(val, float): return str(int(val)) if val == int(val) else str(val)
    return str(val).strip()

def normalize(s):
    s = (s or "").upper().strip()
    s = s.replace('İ','I').replace('Ü','U').replace('Ö','O').replace('Ç','C').replace('Ğ','G').replace('Ş','S')
    return s

def kategori_belirle(kaynak):
    parts = (kaynak or '').split('/')
    dosya = normalize(parts[0]) if len(parts) > 0 else ""
    sekme = normalize(parts[1]) if len(parts) > 1 else ""
    if "IHRACATTAN" in sekme or "GERI DONEN" in sekme:
        return ("ihracat", "İhracattan Geri Dönen Ürün")
    if "TRANSIT" in sekme:
        return ("transit", "Transit")
    if "GEMI" in sekme or "KUMANYA" in sekme:
        return ("gemi", "Gemi Kumanyası")
    if "SERBEST" in dosya or "SERBEST" in sekme:
        return ("serbest", "İthalat")  # Değişiklik 2: Serbest Dolaşım -> İthalat
    return ("diger", sekme if sekme else dosya)

def kelime_eslesme(aranan, urun_adi):
    a = normalize(aranan)
    u = normalize(urun_adi)
    if a == u: return True
    if u.startswith(a + " "): return True
    if u.startswith(a): return True
    return False

def son_guncelleme_tarihi():
    """Veriler klasöründeki en son değiştirilen Excel dosyasının tarihini döndürür."""
    en_son = 0
    if not VERILER.exists():
        return None
    for f in VERILER.iterdir():
        if f.suffix == '.xlsx' and f.is_file():
            mtime = f.stat().st_mtime
            if mtime > en_son:
                en_son = mtime
    if en_son == 0:
        return None
    return datetime.fromtimestamp(en_son).strftime("%d.%m.%Y %H:%M")

def excel_oku():
    kayitlar = []
    if not VERILER.exists():
        print(f"UYARI: {VERILER} klasörü bulunamadı!", flush=True)
        return kayitlar
    for f in VERILER.iterdir():
        if f.suffix != '.xlsx': continue
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            for sekme in wb.sheetnames:
                ws = wb[sekme]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2: continue
                s_norm = normalize(sekme)
                # Varsayılan indeksler (SERBEST DOLAŞIM, TRANSİT, GEMİ KUMANYASI için)
                urun_idx = 8; sebep_idx = 12; karar_idx = 13; ulke_idx = 4
                tarih_idx = 3; kurum_idx = 2; red_sebep_idx = None

                if "IHRACATTAN" in s_norm or "GERI DONEN" in s_norm:
                    # İhracattan Geri Dönen: farklı sütun yapısı
                    urun_idx = 10; sebep_idx = 16; karar_idx = 20; ulke_idx = 4
                    tarih_idx = 3; kurum_idx = 2; red_sebep_idx = 21
                elif "TRANSIT" in s_norm:
                    urun_idx = 8; sebep_idx = 12; karar_idx = 13; ulke_idx = 4
                    tarih_idx = 3; kurum_idx = 2
                elif "GEMI" in s_norm or "KUMANYA" in s_norm:
                    urun_idx = 8; sebep_idx = 12; karar_idx = 13; ulke_idx = 4
                    tarih_idx = 3; kurum_idx = 2
                elif "SERBEST" in s_norm:
                    urun_idx = 8; sebep_idx = 12; karar_idx = 13; ulke_idx = 4
                    tarih_idx = 3; kurum_idx = 2

                baslik = rows[1]
                for j, v in enumerate(baslik):
                    if v is None: continue
                    vs = normalize(str(v)).replace(" ","")
                    if "URUNADI" in vs: urun_idx = j
                    if "UYGUNSUZLUKSEBEBI" in vs: sebep_idx = j
                    if "GERIDONUSSEBEBI" in vs: sebep_idx = j
                    if vs == "KARAR": karar_idx = j
                    if "KONTROLSONUCU" in vs and ("UYGUN" in vs or "RED" in vs): karar_idx = j
                    if ("ULKE" in vs) and ("GONDERICI" in vs or "ITHALATCI" in vs or "GONDEREN" in vs): ulke_idx = j
                    if "BILDIRIMTARIHI" in vs: tarih_idx = j
                    if "BILDIRIMIYAPANKURUM" in vs or "BILDIRIMIYAPAN" in vs: kurum_idx = j
                    if "REDDEDILMESEBEBI" in vs or "REDDEDILENSEVKIYAT" in vs: red_sebep_idx = j

                for row in rows[2:]:
                    if not row or all(v is None for v in row): continue
                    urun = h(row[urun_idx]) if urun_idx < len(row) else ""
                    sebep = h(row[sebep_idx]) if sebep_idx < len(row) else ""
                    karar = h(row[karar_idx]) if karar_idx < len(row) else ""
                    trh = h(row[tarih_idx]) if tarih_idx < len(row) else ""
                    ulke = h(row[ulke_idx]) if ulke_idx < len(row) else ""
                    kurum = h(row[kurum_idx]) if kurum_idx < len(row) else ""
                    red_sebep = h(row[red_sebep_idx]) if red_sebep_idx is not None and red_sebep_idx < len(row) else ""
                    kaynak = f.name + "/" + sekme
                    if urun and urun != "None":
                        kat_kod, kat_ad = kategori_belirle(kaynak)
                        kr = normalize(karar)

                        # Değişiklik 3: Sadece İhracattan Geri Dönen için red/uygun kontrolü
                        if kat_kod == "ihracat":
                            red_kel = ["RED","MAHRECE IADE","IMHA","ULKEYE GIRIS","IZIN VERILMEMIS","UYGUN DEGIL"]
                            is_red = any(r in kr for r in red_kel) if karar else True
                            # Karar sütununda "UYGUN" varsa uygun kabul et
                            if "UYGUN" in kr and "DEGIL" not in kr and "UYGUNSUZ" not in kr:
                                is_red = False
                        else:
                            # Diğer kategorilerde (İthalat, Transit, Gemi Kumanyası) red/uygun ayrımı yok
                            # Hepsi uygunsuzluk kaydı olarak işaretlenir
                            is_red = True

                        kayitlar.append({
                            "urun": urun,
                            "urun_norm": normalize(urun),
                            "tarih": trh,
                            "sebep": sebep if sebep else "-",
                            "karar": karar if karar else "-",
                            "ulke": ulke,
                            "ulke_norm": normalize(ulke),
                            "kurum": kurum,
                            "red_sebep": red_sebep,
                            "kaynak": kaynak,
                            "kategori_kod": kat_kod,
                            "kategori_ad": kat_ad,
                            "red_mi": is_red
                        })
            wb.close()
        except Exception as e:
            import traceback; traceback.print_exc()
    return kayitlar

print("📂 Excel dosyaları okunuyor...", flush=True)
KAYITLAR = excel_oku()
print(f"✅ {len(KAYITLAR)} kayıt yüklendi", flush=True)

@app.route('/')
def index():
    return send_from_directory(PUBLIC, 'index.html')

@app.route('/<path:path>')
def static_dosya(path):
    return send_from_directory(PUBLIC, path)

@app.route('/api/sorgula', methods=['POST'])
def sorgula():
    try:
        data = request.get_json() or {}
        urun = normalize(data.get('urun_adi',''))
        ulke = normalize(data.get('mensei_ulke',''))
        sonuc = []
        for k in KAYITLAR:
            eslesme = True
            if urun: eslesme = eslesme and kelime_eslesme(urun, k["urun"])
            if ulke: eslesme = eslesme and (ulke in k["ulke_norm"] or k["ulke_norm"] in ulke)
            if eslesme: sonuc.append(k)
        sonuc.sort(key=lambda x: (0 if x["red_mi"] else 1), reverse=True)
        return jsonify({"toplam":len(sonuc),"kayitlar":sonuc[:200],"sorgu":data})
    except Exception as e:
        return jsonify({"hata":str(e)}),500

@app.route('/api/son-guncelleme', methods=['GET'])
def son_guncelleme():
    try:
        tarih = son_guncelleme_tarihi()
        return jsonify({"tarih": tarih, "toplam": len(KAYITLAR)})
    except Exception as e:
        return jsonify({"hata": str(e)}), 500

@app.route('/api/kayitlar', methods=['GET'])
def kayitlar():
    try:
        return jsonify({"toplam":len(KAYITLAR),"kayitlar":KAYITLAR[:100]})
    except Exception as e:
        return jsonify({"hata":str(e)}),500

if __name__ == '__main__':
    print("="*60)
    print("İskenderun Limanı VSKN - Ürün Takip Sistemi")
    print(f"\n⚡ Local sunucu: http://127.0.0.1:5000")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
